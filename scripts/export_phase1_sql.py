#!/usr/bin/env python
"""Export Health Assistant OS phase-1 workbook data as PostgreSQL SQL.

The generated SQL contains personal data from the workbook, but it does not
store the old plain PIN values. PINs are converted to PBKDF2-SHA256 hashes
before being written to SQL.
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import json
import re
import secrets
import sys
from pathlib import Path
from typing import Any, Iterable

import openpyxl


PBKDF2_ITERATIONS = 210_000


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("'", "").strip()


def boolish(value: Any, default: bool = False) -> bool:
    text = clean(value).lower()
    if value is True or text in {"true", "1", "yes", "y", "on", "active", "ใช่"}:
        return True
    if value is False or text in {"false", "0", "no", "n", "off", "inactive", "ไม่"}:
        return False
    if text == "":
        return default
    return default


def sql_string(value: Any) -> str:
    text = clean(value)
    if text == "":
        return "null"
    return "'" + text.replace("'", "''") + "'"


def sql_json(value: Any, fallback: Any) -> str:
    if value in (None, ""):
        parsed = fallback
    elif isinstance(value, (dict, list)):
        parsed = value
    else:
        try:
            parsed = json.loads(str(value))
        except Exception:
            parsed = fallback
    return "'" + json.dumps(parsed, ensure_ascii=False).replace("'", "''") + "'::jsonb"


def sql_bool(value: Any, default: bool = False) -> str:
    return "true" if boolish(value, default) else "false"


def sql_timestamp(value: Any, default_now: bool = False) -> str:
    if value in (None, ""):
        return "now()" if default_now else "null"
    if isinstance(value, dt.datetime):
        return "'" + value.isoformat(sep=" ", timespec="seconds") + "+07:00'::timestamptz"
    if isinstance(value, dt.date):
        return "'" + dt.datetime(value.year, value.month, value.day).isoformat(sep=" ", timespec="seconds") + "+07:00'::timestamptz"
    text = clean(value)
    return "'" + text.replace("'", "''") + "'::timestamptz" if text else ("now()" if default_now else "null")


def pbkdf2_hash(pin: Any) -> str:
    pin_text = clean(pin)
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", pin_text.encode("utf-8"), salt.encode("utf-8"), PBKDF2_ITERATIONS)
    return "pbkdf2_sha256${}${}${}".format(PBKDF2_ITERATIONS, salt, base64.b64encode(digest).decode("ascii"))


def worksheet_rows(workbook: openpyxl.Workbook, sheet_name: str) -> tuple[list[str], list[list[Any]]]:
    if sheet_name not in workbook.sheetnames:
        return [], []
    ws = workbook[sheet_name]
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[list[Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(value not in (None, "") for value in row):
            rows.append(row)
    return headers, rows


def cell(row: list[Any], headers: list[str], name: str, fallback_index: int | None = None) -> Any:
    if name in headers:
        idx = headers.index(name)
        return row[idx] if idx < len(row) else ""
    if fallback_index is not None and fallback_index < len(row):
        return row[fallback_index]
    return ""


def account_status(value: Any) -> str:
    return clean(value) or "Active"


def email_notify(value: Any) -> bool:
    text = clean(value).lower()
    return text not in {"off", "false", "0", "no", "none"}


def notification_module(type_value: Any) -> str:
    text = clean(type_value).lower()
    if re.search(r"it booking|booking|zoom|room", text) or "จอง" in text:
        return "itBooking"
    if re.search(r"it repair|repair|helpdesk", text) or "ซ่อม" in text:
        return "itRepair"
    if "license" in text or "asset" in text:
        return "itAsset"
    if re.search(r"e-?meeting|rsvp|meeting action", text):
        return "eMeeting"
    if "chat" in text:
        return "helpChat"
    if "approval" in text or "schedule" in text or "อนุมัติ" in text or "ตาราง" in text:
        return "schedule"
    if "pin" in text or "user" in text or "invite" in text:
        return "userAdmin"
    if "system" in text or "backup" in text or "permission" in text or "database" in text:
        return "systemTools"
    return "system"


def notification_action(module_name: str) -> str:
    return {
        "itBooking": "openItBooking",
        "itRepair": "openITRepair",
        "itAsset": "openITAsset",
        "eMeeting": "openEMeeting",
        "helpChat": "openHelpChat",
        "schedule": "openSchedule",
        "userAdmin": "openUserAdmin",
        "systemTools": "openSystemTools",
    }.get(module_name, "openNotification")


def profile_role(role: Any) -> str:
    text = clean(role) or "Member"
    lower = text.lower()
    if lower in {"head", "department_head"} or "หัวหน้า" in text:
        return "Head"
    if lower in {"acting head", "acting_head"}:
        return "Acting Head"
    if lower == "coordinator" or "ประสาน" in text:
        return "Coordinator"
    return text


def collect_departments(workbook: openpyxl.Workbook) -> list[str]:
    departments: set[str] = set()
    for sheet, column in [
        ("Users", "Department"),
        ("UserProfiles", "Department"),
        ("ITBookings", "Department"),
        ("DailyReports", "Department"),
        ("Schedules", "Owner Department"),
    ]:
        headers, rows = worksheet_rows(workbook, sheet)
        for row in rows:
            value = clean(cell(row, headers, column))
            if value:
                departments.add(value)
    return sorted(departments)


def emit_values_insert(table: str, columns: Iterable[str], values: list[str], conflict: str, updates: Iterable[str] | None = None) -> str:
    cols = list(columns)
    update_cols = list(updates or cols)
    update_sql = ", ".join(f"{col}=excluded.{col}" for col in update_cols if col not in conflict.split(","))
    if update_sql:
        action = f"do update set {update_sql}"
    else:
        action = "do nothing"
    return (
        f"insert into public.{table} ({', '.join(cols)})\n"
        f"values ({', '.join(values)})\n"
        f"on conflict ({conflict}) {action};"
    )


def generate_sql(workbook_path: Path) -> tuple[str, dict[str, Any]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    statements: list[str] = [
        "-- Health Assistant OS phase-1 data import",
        "-- Generated from workbook with PIN values hashed as PBKDF2-SHA256.",
        "begin;",
    ]
    summary: dict[str, Any] = {"workbook": str(workbook_path), "sheets": {}, "warnings": []}

    departments = collect_departments(workbook)
    for order, department in enumerate(departments, start=1):
        statements.append(
            emit_values_insert(
                "departments",
                ["department_name", "sort_order", "is_active"],
                [sql_string(department), str(order), "true"],
                "department_name",
                ["sort_order", "is_active"],
            )
        )
    summary["departments"] = len(departments)

    headers, rows = worksheet_rows(workbook, "Users")
    summary["sheets"]["Users"] = len(rows)
    for row in rows:
        phone = clean(cell(row, headers, "Phone (Username)", 0))
        if not phone:
            summary["warnings"].append("Skipped user row with blank phone")
            continue
        department = clean(cell(row, headers, "Department", 4))
        status = account_status(cell(row, headers, "Account Status", 10))
        if clean(cell(row, headers, "Account Status", 10)) == "":
            summary["warnings"].append(f"User {phone} had blank Account Status; imported as Active")
        values = [
            sql_string(phone),
            sql_string(pbkdf2_hash(cell(row, headers, "PIN (Password)", 1))),
            "'pbkdf2_sha256'",
            sql_string(cell(row, headers, "Full Name", 2)),
            sql_string(cell(row, headers, "Position", 3)),
            f"(select department_id from public.departments where department_name={sql_string(department)} limit 1)",
            sql_string(department),
            sql_string(cell(row, headers, "Email", 5)),
            sql_string(cell(row, headers, "PDPA Status", 6)),
            sql_string(clean(cell(row, headers, "Role", 7)) or "User"),
            sql_string(status),
            "true" if email_notify(cell(row, headers, "Email Notify", 9)) else "false",
            sql_json(cell(row, headers, "Email Notify Preferences JSON"), {}),
            sql_string(cell(row, headers, "Signature Data")),
            sql_json(cell(row, headers, "Signature List JSON"), []),
            sql_timestamp(cell(row, headers, "Approved At")),
            sql_string(cell(row, headers, "Approved By")),
            sql_string(cell(row, headers, "Invite Token")),
            sql_timestamp(cell(row, headers, "Last Login", 11)),
            sql_timestamp(cell(row, headers, "Timestamp", 8)),
        ]
        statements.append(
            emit_values_insert(
                "users",
                [
                    "phone", "pin_hash", "pin_hash_algorithm", "full_name", "position", "department_id",
                    "department_name", "email", "pdpa_status", "account_role", "account_status",
                    "email_notify", "email_notify_preferences", "signature_data", "signature_list",
                    "approved_at", "approved_by_phone", "invite_token", "last_login_at", "legacy_created_at",
                ],
                values,
                "phone",
                [
                    "pin_hash", "pin_hash_algorithm", "full_name", "position", "department_id",
                    "department_name", "email", "pdpa_status", "account_role", "account_status",
                    "email_notify", "email_notify_preferences", "signature_data", "signature_list",
                    "approved_at", "approved_by_phone", "invite_token", "last_login_at", "legacy_created_at",
                ],
            )
        )

    headers, rows = worksheet_rows(workbook, "UserProfiles")
    summary["sheets"]["UserProfiles"] = len(rows)
    for row in rows:
        phone = clean(cell(row, headers, "User Phone", 1))
        department = clean(cell(row, headers, "Department", 3))
        profile_id = clean(cell(row, headers, "Profile ID", 0))
        if not phone or not profile_id:
            summary["warnings"].append(f"Skipped profile row with blank phone/profile ID: {profile_id or '-'}")
            continue
        values = [
            sql_string(profile_id),
            f"(select user_id from public.users where phone={sql_string(phone)} limit 1)",
            sql_string(phone),
            sql_string(cell(row, headers, "Full Name Snapshot", 2)),
            f"(select department_id from public.departments where department_name={sql_string(department)} limit 1)",
            sql_string(department),
            sql_string(cell(row, headers, "Position", 4)),
            sql_string(profile_role(cell(row, headers, "Profile Role", 5))),
            sql_bool(cell(row, headers, "Is Primary", 6)),
            sql_bool(cell(row, headers, "Is Active", 7), True),
            sql_bool(cell(row, headers, "Can Approve", 8)),
            sql_string(cell(row, headers, "Signature ID", 9)),
            sql_string(cell(row, headers, "Notify Email", 10)),
            sql_string(cell(row, headers, "Notify Telegram", 11)),
            sql_string(cell(row, headers, "Notes", 15)),
            sql_string(cell(row, headers, "Updated By", 14)),
            sql_timestamp(cell(row, headers, "Created At", 12), True),
            sql_timestamp(cell(row, headers, "Updated At", 13), True),
        ]
        statements.append(
            emit_values_insert(
                "user_profiles",
                [
                    "profile_id", "user_id", "user_phone", "full_name_snapshot", "department_id",
                    "department_name", "position", "profile_role", "is_primary", "is_active",
                    "can_approve", "signature_id", "notify_email", "notify_telegram", "notes",
                    "updated_by_phone", "created_at", "updated_at",
                ],
                values,
                "profile_id",
                [
                    "user_id", "user_phone", "full_name_snapshot", "department_id", "department_name",
                    "position", "profile_role", "is_primary", "is_active", "can_approve", "signature_id",
                    "notify_email", "notify_telegram", "notes", "updated_by_phone", "updated_at",
                ],
            )
        )

    user_headers, user_rows = worksheet_rows(workbook, "Users")
    for row in user_rows:
        phone = clean(cell(row, user_headers, "Phone (Username)", 0))
        active_profile_id = clean(cell(row, user_headers, "Active Profile ID"))
        if phone and active_profile_id:
            statements.append(
                "update public.users set active_profile_id="
                + sql_string(active_profile_id)
                + " where phone="
                + sql_string(phone)
                + ";"
            )

    headers, rows = worksheet_rows(workbook, "AutoLoginDevices")
    summary["sheets"]["AutoLoginDevices"] = len(rows)
    for row in rows:
        phone = clean(cell(row, headers, "User Phone", 1))
        device_id = clean(cell(row, headers, "Device ID", 0))
        token_hash = clean(cell(row, headers, "Token Hash", 3))
        if not phone or not device_id or not token_hash:
            summary["warnings"].append(f"Skipped auto-login device row with missing key: {device_id or '-'}")
            continue
        statements.append(
            emit_values_insert(
                "auto_login_devices",
                [
                    "device_id", "user_id", "user_phone", "device_label", "token_hash", "user_agent",
                    "created_at", "last_used_at", "expires_at", "is_active", "revoked_at",
                    "revoked_by_phone", "notes",
                ],
                [
                    sql_string(device_id),
                    f"(select user_id from public.users where phone={sql_string(phone)} limit 1)",
                    sql_string(phone),
                    sql_string(cell(row, headers, "Device Label", 2)),
                    sql_string(token_hash),
                    sql_string(cell(row, headers, "User Agent", 4)),
                    sql_timestamp(cell(row, headers, "Created At", 5), True),
                    sql_timestamp(cell(row, headers, "Last Used At", 6)),
                    sql_timestamp(cell(row, headers, "Expires At", 7)),
                    sql_bool(cell(row, headers, "Is Active", 8), True),
                    sql_timestamp(cell(row, headers, "Revoked At", 9)),
                    sql_string(cell(row, headers, "Revoked By", 10)),
                    sql_string(cell(row, headers, "Notes", 11)),
                ],
                "user_id, device_id",
                [
                    "user_phone", "device_label", "token_hash", "user_agent", "last_used_at",
                    "expires_at", "is_active", "revoked_at", "revoked_by_phone", "notes", "updated_at",
                ],
            )
        )

    headers, rows = worksheet_rows(workbook, "Notifications")
    summary["sheets"]["Notifications"] = len(rows)
    missing_module_action = 0
    for row in rows:
        phone = clean(cell(row, headers, "Target Phone", 1))
        notif_id = clean(cell(row, headers, "Notif ID", 0))
        type_value = clean(cell(row, headers, "Type", 5))
        module_value = clean(cell(row, headers, "Module"))
        action_value = clean(cell(row, headers, "Action"))
        if not module_value:
            module_value = notification_module(type_value)
            missing_module_action += 1
        if not action_value:
            action_value = notification_action(module_value)
        if not phone or not notif_id:
            summary["warnings"].append(f"Skipped notification row with missing key: {notif_id or '-'}")
            continue
        created_by_phone = clean(cell(row, headers, "Created By"))
        statements.append(
            emit_values_insert(
                "notifications",
                [
                    "notification_id", "target_user_id", "target_phone", "message", "is_read", "type",
                    "entity_id", "priority", "module", "action", "email_sent_at", "email_status",
                    "created_by_user_id", "created_by_phone", "read_at", "created_at",
                ],
                [
                    sql_string(notif_id),
                    f"(select user_id from public.users where phone={sql_string(phone)} limit 1)",
                    sql_string(phone),
                    sql_string(cell(row, headers, "Message", 2)),
                    sql_bool(cell(row, headers, "Is Read", 3)),
                    sql_string(type_value),
                    sql_string(cell(row, headers, "Entity ID", 6)),
                    sql_string(clean(cell(row, headers, "Priority", 7)) or "Normal"),
                    sql_string(module_value),
                    sql_string(action_value),
                    sql_timestamp(cell(row, headers, "Email Sent At", 8)),
                    sql_string(cell(row, headers, "Email Status", 9)),
                    f"(select user_id from public.users where phone={sql_string(created_by_phone)} limit 1)",
                    sql_string(created_by_phone),
                    sql_timestamp(cell(row, headers, "Read At")),
                    sql_timestamp(cell(row, headers, "Timestamp", 4), True),
                ],
                "notification_id",
                [
                    "target_user_id", "target_phone", "message", "is_read", "type", "entity_id",
                    "priority", "module", "action", "email_sent_at", "email_status", "created_by_user_id",
                    "created_by_phone", "read_at", "updated_at",
                ],
            )
        )
    if missing_module_action:
        summary["warnings"].append(f"Derived Module/Action for {missing_module_action} old notification rows")

    headers, rows = worksheet_rows(workbook, "AuditLogs")
    summary["sheets"]["AuditLogs"] = len(rows)
    for row in rows:
        audit_id = clean(cell(row, headers, "Audit ID", 0))
        actor_phone = clean(cell(row, headers, "Actor Phone", 2))
        if not audit_id:
            summary["warnings"].append("Skipped audit row with blank Audit ID")
            continue
        statements.append(
            emit_values_insert(
                "audit_logs",
                [
                    "audit_id", "actor_user_id", "actor_phone", "actor_name", "action", "entity_type",
                    "entity_id", "summary", "before_json", "after_json", "created_at",
                ],
                [
                    sql_string(audit_id),
                    f"(select user_id from public.users where phone={sql_string(actor_phone)} limit 1)",
                    sql_string(actor_phone),
                    sql_string(cell(row, headers, "Actor Name", 3)),
                    sql_string(cell(row, headers, "Action", 4)),
                    sql_string(cell(row, headers, "Entity Type", 5)),
                    sql_string(cell(row, headers, "Entity ID", 6)),
                    sql_string(cell(row, headers, "Summary", 7)),
                    sql_json(cell(row, headers, "Before JSON", 8), None),
                    sql_json(cell(row, headers, "After JSON", 9), None),
                    sql_timestamp(cell(row, headers, "Timestamp", 1), True),
                ],
                "audit_id",
                [],
            )
        )

    headers, rows = worksheet_rows(workbook, "Resources")
    summary["sheets"]["Resources"] = len(rows)
    for row in rows:
        resource_id = clean(cell(row, headers, "Resource ID", 0))
        if not resource_id:
            continue
        statements.append(
            emit_values_insert(
                "resources",
                ["resource_id", "resource_name", "resource_type", "is_active"],
                [sql_string(resource_id), sql_string(cell(row, headers, "Resource Name", 1)), sql_string(cell(row, headers, "Type", 2)), "true"],
                "resource_id",
                ["resource_name", "resource_type", "is_active"],
            )
        )

    headers, rows = worksheet_rows(workbook, "Settings")
    summary["sheets"]["Settings"] = len(rows)
    for row in rows:
        key = clean(cell(row, headers, "Key", 0))
        if not key:
            continue
        statements.append(
            emit_values_insert(
                "settings",
                ["key", "value", "description", "updated_at"],
                [sql_string(key), sql_json(cell(row, headers, "Value", 1), clean(cell(row, headers, "Value", 1))), sql_string(cell(row, headers, "Description", 2)), sql_timestamp(cell(row, headers, "Updated At", 3), True)],
                "key",
                ["value", "description", "updated_at"],
            )
        )

    statements.extend(
        [
            "commit;",
            "",
            "-- Validation queries",
            "select 'users' as table_name, count(*) from public.users",
            "union all select 'user_profiles', count(*) from public.user_profiles",
            "union all select 'auto_login_devices', count(*) from public.auto_login_devices",
            "union all select 'notifications', count(*) from public.notifications",
            "union all select 'audit_logs', count(*) from public.audit_logs;",
        ]
    )
    return "\n\n".join(statements), summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Export phase-1 PostgreSQL seed SQL from a Health Assistant OS workbook.")
    parser.add_argument("--workbook", default="AI Assistant OS (29-5-69).xlsx", help="Path to the latest workbook")
    parser.add_argument("--output", help="Output SQL path. Omit to print SQL to stdout.")
    parser.add_argument("--dry-run", action="store_true", help="Only validate workbook and print summary; do not emit SQL.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    sql, summary = generate_sql(workbook_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2), file=sys.stderr)
    if args.dry_run:
        return 0

    if args.output:
        output = Path(args.output).resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(sql, encoding="utf-8")
        print(f"Wrote SQL import file: {output}", file=sys.stderr)
    else:
        print(sql)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
